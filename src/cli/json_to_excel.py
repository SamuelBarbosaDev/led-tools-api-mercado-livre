import argparse
import json
from pathlib import Path

import pandas as pd


DEFAULT_COLUMNS = [
    "id",
    "title",
    "category",          # nome da categoria (se você adicionou)
    "price",
    "sold_quantity",
    "available_quantity",
    "free_shipping",
    "logistic_type",
    "permalink",
    "picture_url",
]


def autosize_worksheet(ws, df: pd.DataFrame, max_width: int = 60) -> None:
    """Ajuste simples de largura das colunas baseado no conteúdo."""
    for idx, col in enumerate(df.columns, start=1):
        # largura: max(tamanho do header, tamanho do maior valor em string) + 2
        series_as_str = df[col].astype(str).fillna("")
        best = max(len(str(col)), series_as_str.map(len).max() if not series_as_str.empty else 0)
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(best + 2, max_width)


def apply_excel_formatting(path: Path, sheet_name: str = "items") -> None:
    """Aplica formatações usando openpyxl (já vem junto com pandas engine)."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    ws = wb[sheet_name]

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Auto-filtro
    ws.auto_filter.ref = ws.dimensions

    # Estilo do cabeçalho
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    # Formatar coluna de preço como moeda (R$)
    # (se existir e se for número)
    header = [c.value for c in ws[1]]
    if "price" in header:
        price_col = header.index("price") + 1
        col_letter = get_column_letter(price_col)
        for row in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row}"].number_format = 'R$ #,##0.00'

    # Alinhar links/permalink para não ficar feio (opcional)
    for col_name in ("permalink", "picture_url"):
        if col_name in header:
            col = header.index(col_name) + 1
            col_letter = get_column_letter(col)
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].alignment = Alignment(wrap_text=False)

    wb.save(path)


def main() -> None:
    p = argparse.ArgumentParser(description="Converte JSON (lista de itens) em Excel (.xlsx).")
    p.add_argument("json_file", help="Caminho do JSON de entrada (ex: items.json)")
    p.add_argument("-o", "--out", default=None, help="Caminho do Excel de saída (ex: items.xlsx)")
    p.add_argument(
        "--columns",
        nargs="*",
        default=DEFAULT_COLUMNS,
        help="Lista de colunas (em ordem) a exportar. Colunas inexistentes serão ignoradas.",
    )
    args = p.parse_args()

    json_path = Path(args.json_file)
    if not json_path.exists():
        raise SystemExit(f"Arquivo não encontrado: {json_path}")

    out_path = Path(args.out) if args.out else json_path.with_suffix(".xlsx")

    # Ler JSON
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise SystemExit("JSON inválido: esperado uma lista de objetos (list[dict]).")

    df = pd.DataFrame(data)

    # Ordenar/selecionar colunas (ignora as que não existem)
    cols = [c for c in args.columns if c in df.columns]
    other_cols = [c for c in df.columns if c not in cols]
    df = df[cols + other_cols] if cols else df

    # Garantir tipos numéricos onde faz sentido
    for col in ("price", "sold_quantity", "available_quantity"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Exportar
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="items")
        ws = writer.book["items"]
        autosize_worksheet(ws, df)

    # Pós-formatação (moeda, freeze, filtro, etc.)
    apply_excel_formatting(out_path, sheet_name="items")

    print(f"OK: Excel gerado em {out_path}")


if __name__ == "__main__":
    main()
