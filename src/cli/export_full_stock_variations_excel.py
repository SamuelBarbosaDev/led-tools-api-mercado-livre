#!/usr/bin/env python3
# src/cli/export_full_stock_variations_excel.py

from __future__ import annotations

from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ledtools_ml.config import ML_SELLER_USER_ID
from ledtools_ml.ml import list_item_ids_public, get_item, get_user_product_stock


def utc_today() -> str:
    return datetime.now(timezone.utc).date().isoformat()


def pick_location_qty(stock_payload: Dict[str, Any], loc_type: str) -> int:
    """
    stock_payload esperado:
      {"locations":[{"type":"meli_facility","quantity":5}, ...]}
    """
    locs = stock_payload.get("locations") or []
    if not isinstance(locs, list):
        return 0
    for loc in locs:
        if isinstance(loc, dict) and loc.get("type") == loc_type:
            q = loc.get("quantity")
            return int(q) if isinstance(q, (int, float)) else 0
    return 0


def variation_attrs_text(variation: Dict[str, Any]) -> str:
    """
    Converte atributos de variação em um texto "Cor=Preto; Tamanho=G".
    Campo comum: variation["attribute_combinations"] (lista de {name, value_name})
    """
    combos = variation.get("attribute_combinations") or []
    parts: List[str] = []
    if isinstance(combos, list):
        for a in combos:
            if not isinstance(a, dict):
                continue
            name = a.get("name")
            val = a.get("value_name") or a.get("value_id")
            if name and val:
                parts.append(f"{name}={val}")
    return "; ".join(parts)


def extract_variations(item: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Retorna lista de dicts no formato:
      {
        "variation_id": ...,
        "user_product_id": ...,
        "attrs": "Cor=...",
      }
    """
    variations = item.get("variations") or []
    out: List[Dict[str, Any]] = []

    if not isinstance(variations, list) or not variations:
        return out

    for v in variations:
        if not isinstance(v, dict):
            continue
        upid = v.get("user_product_id")
        vid = v.get("id")
        if not upid or not vid:
            # Sem user_product_id não dá pra consultar stock por user-products
            # (ainda assim você pode exportar como None se quiser)
            continue

        out.append(
            {
                "variation_id": str(vid),
                "user_product_id": str(upid),
                "attrs": variation_attrs_text(v),
            }
        )
    return out


def autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 10
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 70)


def main() -> None:
    seller_id = ML_SELLER_USER_ID
    if not seller_id:
        raise SystemExit("Defina ML_SELLER_USER_ID no .env (ex.: 570565928)")

    today = utc_today()
    item_ids = list_item_ids_public(seller_id)

    rows: List[Dict[str, Any]] = []

    for item_id in item_ids:
        item = get_item(item_id)

        title = item.get("title")
        category_id = item.get("category_id")
        logistic_type = (item.get("shipping") or {}).get("logistic_type")

        variations = extract_variations(item)

        # Se o item não tiver variações (ou não tiver user_product_id por variação),
        # você pode optar por:
        #  - pular
        #  - ou exportar uma linha “sem variação”
        # Aqui vou exportar uma linha “sem variação” apenas para você saber que existe.
        if not variations:
            rows.append(
                {
                    "date": today,
                    "item_id": item.get("id"),
                    "variation_id": None,
                    "user_product_id": item.get("user_product_id"),
                    "title": title,
                    "category_id": category_id,
                    "variation_attributes": None,
                    "full_stock": None,
                    "seller_stock": None,
                    "logistic_type": logistic_type,
                }
            )
            continue

        # 1 linha por variação
        for v in variations:
            stock = get_user_product_stock(v["user_product_id"])
            full_qty = pick_location_qty(stock, "meli_facility")      # FULL
            seller_qty = pick_location_qty(stock, "selling_address")  # estoque vendedor (se vier)

            rows.append(
                {
                    "date": today,
                    "item_id": item.get("id"),
                    "variation_id": v["variation_id"],
                    "user_product_id": v["user_product_id"],
                    "title": title,
                    "category_id": category_id,
                    "variation_attributes": v["attrs"],
                    "full_stock": full_qty,
                    "seller_stock": seller_qty,
                    "logistic_type": logistic_type,
                }
            )

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Full Stock (Variations)"

    headers = [
        "date",
        "item_id",
        "variation_id",
        "user_product_id",
        "title",
        "category_id",
        "variation_attributes",
        "full_stock",
        "seller_stock",
        "logistic_type",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for r in rows:
        ws.append([r.get(h) for h in headers])

    autosize(ws)

    out = "ml_full_stock_variations.xlsx"
    wb.save(out)
    print(f"OK: {out} com {len(rows)} linhas")


if __name__ == "__main__":
    main()