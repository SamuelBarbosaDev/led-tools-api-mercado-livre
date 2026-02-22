# src/cli/export_full_stock_excel.py
from __future__ import annotations

from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ledtools_ml.config import ML_SELLER_USER_ID
from ledtools_ml.ml import list_item_ids_public, get_item
from ledtools_ml.http import request_json

BASE = "https://api.mercadolibre.com"

def utc_today() -> str:
    return datetime.now(timezone.utc).date().isoformat()

def find_user_product_ids(item: Dict[str, Any]) -> List[str]:
    ids: List[str] = []

    # 1) Tenta no item
    upid = item.get("user_product_id")
    if isinstance(upid, str) and upid:
        ids.append(upid)

    # 2) Tenta nas variações
    variations = item.get("variations") or []
    if isinstance(variations, list):
        for v in variations:
            if isinstance(v, dict):
                v_upid = v.get("user_product_id")
                if isinstance(v_upid, str) and v_upid:
                    ids.append(v_upid)

    # Remove duplicados preservando ordem
    seen = set()
    out = []
    for x in ids:
        if x in seen:
            continue
        seen.add(x)
        out.append(x)
    return out

def get_user_product_stock(user_product_id: str) -> Dict[str, Any]:
    return request_json("GET", f"{BASE}/user-products/{user_product_id}/stock")

def pick_location_qty(stock_payload: Dict[str, Any], loc_type: str) -> int:
    locs = stock_payload.get("locations") or []
    if not isinstance(locs, list):
        return 0
    for loc in locs:
        if isinstance(loc, dict) and loc.get("type") == loc_type:
            q = loc.get("quantity")
            return int(q) if isinstance(q, (int, float)) else 0
    return 0

def autosize(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 10
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

def main():
    seller_id = ML_SELLER_USER_ID
    if not seller_id:
        raise SystemExit("Defina ML_SELLER_USER_ID no .env")

    item_ids = list_item_ids_public(seller_id)

    rows = []
    today = utc_today()

    for item_id in item_ids:
        item = get_item(item_id)

        user_product_ids = find_user_product_ids(item)
        # Se não tiver user_product_id, ainda dá pra exportar “sem estoque full”
        if not user_product_ids:
            rows.append({
                "date": today,
                "item_id": item.get("id"),
                "title": item.get("title"),
                "category_id": item.get("category_id"),
                "user_product_id": None,
                "full_stock": None,
                "flex_stock": None,
                "logistic_type": (item.get("shipping") or {}).get("logistic_type"),
            })
            continue

        # Se tiver variações, exporta uma linha por user_product_id
        for upid in user_product_ids:
            stock = get_user_product_stock(upid)
            full_qty = pick_location_qty(stock, "meli_facility")      # FULL
            flex_qty = pick_location_qty(stock, "selling_address")    # estoque vendedor (se aplicável)

            rows.append({
                "date": today,
                "item_id": item.get("id"),
                "title": item.get("title"),
                "category_id": item.get("category_id"),
                "user_product_id": upid,
                "full_stock": full_qty,
                "flex_stock": flex_qty,
                "logistic_type": (item.get("shipping") or {}).get("logistic_type"),
            })

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Full Stock"

    headers = ["date","item_id","title","category_id","user_product_id","full_stock","flex_stock","logistic_type"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for r in rows:
        ws.append([r[h] for h in headers])

    autosize(ws)
    out = "ml_full_stock.xlsx"
    wb.save(out)
    print(f"OK: {out} com {len(rows)} linhas")

if __name__ == "__main__":
    main()